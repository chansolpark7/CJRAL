import main
import random
import sys
import numpy as np
import openpyxl
from collections import namedtuple

import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
from matplotlib.widgets import Slider
import matplotlib.ticker as ticker

Data = namedtuple(
    'Data',
    [
        'Vehicle_ID',
        'Route_Order',
        'Destination',
        'Order_Number',
        'Box_ID',
        'Stacking_Order',
        'Lower_Left_X',
        'Lower_Left_Y',
        'Lower_Left_Z',
        'Longitude',
        'Latitude',
        'Box_Width',
        'Box_Length',
        'Box_Height'
    ]
)

class box_viewer_3d:
    def __init__(self, box_list, mode=1):
        self.fig = plt.figure(figsize=(10, 8))
        self.fig.canvas.mpl_disconnect(self.fig.canvas.manager.key_press_handler_id)
        self.fig.canvas.mpl_connect('key_press_event', self.key_callback)
        self.ax = self.fig.add_subplot(111, projection='3d')
        self.ax.set_xlim([0, 16])
        self.ax.set_ylim([0, 28])
        self.ax.set_zlim([0, 18])
        self.ax.invert_yaxis()
        self.ax.set_box_aspect([16, 28, 18])

        self.ax.set_xlabel('X')
        self.ax.set_ylabel('Y')
        self.ax.set_zlabel('Z')
        plt.title('3D Box Visualization')

        plt.subplots_adjust(bottom=0.25)
        self.slider_ax = self.fig.add_axes([0.1, 0.1, 0.8, 0.03])
        self.slider = Slider(self.slider_ax, 'Time', 0, len(box_list), valinit=len(box_list), valstep=1)
        self.slider.on_changed(self.update)

        self.box_list = box_list
        self.mode = mode
        self.lines = []
        self.artists = []

        self.colors = [[random.random() * 0.7 + 0.3 for _ in range(3)] for _ in range(200)]
        self.update(self.slider.val)

    def key_callback(self, event):
        val = int(self.slider.val)
        if event.key == 'right' and val < self.slider.valmax:
            self.slider.set_val(val + 1)
        elif event.key == 'left' and val > self.slider.valmin:
            self.slider.set_val(val - 1)

    def update(self, value):
        if self.mode == 0: # line
            for line in self.lines:
                line.remove()
            self.lines = []

            for position, size in self.box_list[:value]:
                x, y, z = position
                dx, dy, dz = size
                # Draw a 3D box as a rectangular prism
                xx = [x, x+dx, x+dx, x, x]
                yy = [y, y, y+dy, y+dy, y]
                kwargs = {'alpha': 0.5}
                # Bottom face
                self.lines += self.ax.plot3D(xx, yy, [z]*5, color='b', **kwargs)
                # Top face
                self.lines += self.ax.plot3D(xx, yy, [z+dz]*5, color='r', **kwargs)
                # Vertical lines
                for i in range(4):
                    self.lines += self.ax.plot3D([xx[i], xx[i]], [yy[i], yy[i]], [z, z+dz], color='g', **kwargs)
        elif self.mode == 1: # face
            for artist in self.artists:
                artist.remove()
            self.artists = []

            for index, (position, size) in enumerate(self.box_list[:value]):
                x, y, z = position
                dx, dy, dz = size

                # 8개 꼭짓점
                corners = [
                    [x, y, z],
                    [x + dx, y, z],
                    [x + dx, y + dy, z],
                    [x, y + dy, z],
                    [x, y, z + dz],
                    [x + dx, y, z + dz],
                    [x + dx, y + dy, z + dz],
                    [x, y + dy, z + dz],
                ]

                # 6개 면 (각각 꼭짓점 4개)
                faces = [
                    [corners[0], corners[1], corners[2], corners[3]],  # bottom
                    [corners[4], corners[5], corners[6], corners[7]],  # top
                    [corners[0], corners[1], corners[5], corners[4]],  # front
                    [corners[2], corners[3], corners[7], corners[6]],  # back
                    [corners[1], corners[2], corners[6], corners[5]],  # right
                    [corners[3], corners[0], corners[4], corners[7]],  # left
                ]

                face_color = self.colors[index]
                box = Poly3DCollection(faces, facecolors=face_color, edgecolors='k', linewidths=0.5, alpha=0.5)
                self.ax.add_collection3d(box)
                self.artists.append(box)

        plt.draw()

    def show(self):
        plt.show()

class box_viewer_2d:
    def __init__(self, box_list):
        self.fig = plt.figure(figsize=(10, 8))
        self.fig.canvas.mpl_disconnect(self.fig.canvas.manager.key_press_handler_id)
        self.fig.canvas.mpl_connect('key_press_event', self.key_callback)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_xlim([0, 16])
        self.ax.set_ylim([0, 28])
        self.ax.invert_yaxis()
        plt.gca().set_aspect('equal', adjustable='box')

        self.ax.set_xlabel('X')
        self.ax.set_ylabel('Y')
        plt.title('2D Box Visualization')

        plt.subplots_adjust(bottom=0.25)
        self.slider_ax = self.fig.add_axes([0.1, 0.1, 0.8, 0.03])
        self.slider = Slider(self.slider_ax, 'Time', 0, len(box_list), valinit=len(box_list), valstep=1)
        self.slider.on_changed(self.update)

        self.box_list = box_list
        self.lines = []

        self.colors = [[random.random() * 0.7 + 0.3 for _ in range(3)] for _ in range(200)]
        self.update(self.slider.val)

    def key_callback(self, event):
        val = int(self.slider.val)
        if event.key == 'right' and val < self.slider.valmax:
            self.slider.set_val(val + 1)
        elif event.key == 'left' and val > self.slider.valmin:
            self.slider.set_val(val - 1)

    def update(self, val):
        for line in self.lines:
            line.remove()
        self.lines = []

        for position, size in self.box_list[:val]:
            x, y = position
            dx, dy = size
            xx = [x, x+dx, x+dx, x, x]
            yy = [y, y, y+dy, y+dy, y]
            kwargs = {'alpha': 0.5}
            self.lines += self.ax.plot(xx, yy, color='rg'[(10, 7).index(sum(size))], **kwargs)
        plt.draw()

    def show(self):
        plt.show()

def histogram(data, start=0, end=1, bins=30):
    plt.hist(data, bins=30, color='skyblue', edgecolor='black', range=(start, end))
    plt.title('Histogram')
    plt.xlabel('Value')
    plt.ylabel('Frequency')
    plt.grid(True)
    plt.show()

def graph(**datas):
    for name, data in datas.items():
        plt.plot(data, label=name)
    plt.title('graph')
    plt.xlabel('index')
    plt.ylabel('value')
    plt.legend()
    plt.grid(True)
    plt.show()

def benchmark(result):
    def int_with_commas(x, pos):
        return f'{int(x):,}'

    n = len(result[0])
    data = []
    for i in range(n):
        time = []
        cost = []
        for d in result:
            time.append(d[i][1])
            cost.append(d[i][2])
        data.append((time, cost))

    x = list(range(1, len(result) + 1))

    # 한 화면에 두 그래프 (Running Time & Total Cost)
    plt.figure(figsize=(12, 5))

    # 1. Running Time
    plt.subplot(1, 2, 1)  # (행, 열, 위치)
    for i in range(n):
        name = f'Algorithm {i+1}'
        plt.plot(x, data[i][0], marker='o', label=name)
    plt.title('Running Time Comparison')
    plt.xlabel('Experiment')
    plt.ylabel('Time')
    plt.legend()
    plt.grid(True)
    plt.ylim(bottom=0)
    plt.gca().xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax = plt.gca()
    formatter = ticker.ScalarFormatter(useOffset=False)
    formatter.set_scientific(False)
    ax.yaxis.set_major_formatter(formatter)

    # 2. Total Cost
    plt.subplot(1, 2, 2)
    for i in range(n):
        name = f'Algorithm {i+1}'
        plt.plot(x, data[i][1], marker='o', label=name)
    plt.title('Total Cost Comparison')
    plt.xlabel('Experiment')
    plt.ylabel('Cost')
    plt.legend()
    plt.grid(True)
    plt.ylim(bottom=0)
    plt.gca().xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax = plt.gca()
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(int_with_commas))

    plt.tight_layout()
    plt.show()

def plot_vrp(data_file_name, distance_file_name):
    destinations, name_to_index, index_to_name = main.read_map(data_file_name)
    n = len(destinations)
    OD_matrix = main.read_OD_matrix(n, name_to_index, distance_file_name)
    orders = main.read_orders(n, name_to_index, data_file_name)

    result_file_name = 'Result.xlsx'
    wb = openpyxl.load_workbook(result_file_name)
    wb.sheetnames
    ws = wb[wb.sheetnames[0]]

    datas: list[list[Data]] = []
    for index, row in enumerate(ws.rows):
        if index == 0: continue
        data = Data(*map(lambda x: x.value, row))
        if len(datas) == data.Vehicle_ID:
            datas.append([])
        datas[data.Vehicle_ID].append(data)

    vehicles: list[main.Vehicle] = []
    for data in datas:
        route = []
        for d in data:
            node = name_to_index[d.Destination]
            if not route or route[-1] != node: route.append(node)
        vehicle = main.Vehicle(route, orders)
        for d in data:
            if d.Box_ID == None: continue
            size = (d.Box_Width//10, d.Box_Length//10, d.Box_Height//10)
            position = (d.Lower_Left_X//10, 28-(d.Lower_Left_Y//10 + size[1]), d.Lower_Left_Z//10)
            vehicle.load_box_at(position, size)
        vehicle.calculate_dist(OD_matrix)
        vehicles.append(vehicle)
    plt.figure(figsize=(10, 8))

    # depot 그리기
    plt.scatter(destinations['Depot'].longitude, destinations['Depot'].latitude, c='red', marker='s', s=100, label='Depot')

    # destination 그리기
    for dest_id, pos in list(destinations.items())[1:]:
        x = pos.longitude
        y = pos.latitude
        plt.scatter(x, y, c='blue')
        # plt.text(x, y, str(dest_id), fontsize=9)

    # 차량 경로 그리기
    for vehicle in vehicles:
        x_vals = []
        y_vals = []
        for route_index in vehicle.route:
            destination_id = index_to_name[route_index]
            destination = destinations[destination_id]
            # coords.append((destination.longitude, destination.latitude))
            x_vals.append(destination.longitude)
            y_vals.append(destination.latitude)

        # 색상과 선 굵기 설정
        color = [random.random() for _ in range(3)]  # 랜덤 색
        thickness = max(1, 5 - vehicle.calc_empty_volume()/vehicle.total_volume*5)  # 빈공간 20%마다 1씩 줄어듦

        plt.plot(x_vals, y_vals, color=color, linewidth=thickness,
                 label=f'Vehicle ({vehicle.calc_empty_volume()/vehicle.total_volume*100: .2f}% left)')

    plt.title('Vehicle Routing Problem')
    plt.xlabel('X')
    plt.ylabel('Y')
    plt.legend()
    plt.grid(True)
    plt.show()

if __name__ == "__main__":
    data_file_name = 'Data_Set.json'
    distance_file_name = 'distance-data.txt'
    # data_file_name = 'additional_data.json'
    # distance_file_name = 'additional_distance_data.txt'

    

    plot_vrp(data_file_name, distance_file_name)