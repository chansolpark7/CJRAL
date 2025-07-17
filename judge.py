import openpyxl
from os.path import basename, dirname, join
import os
import time
from collections import namedtuple

import main
import visualize

DEBUG = False
VISUALIZE = False

Data = namedtuple(
    'Data',
    [
        'Vehicle_ID',
        'Route_Order',
        'Destination',
        'Order_Number',
        'Box_ID',
        'Stacking_Order',
        'Lower_Left_X',
        'Lower_Left_Y',
        'Lower_Left_Z',
        'Longitude',
        'Latitude',
        'Box_Width',
        'Box_Length',
        'Box_Height'
    ]
)

Box = namedtuple(
    'Box',
    [
        'Lower_Left_X',
        'Lower_Left_Y',
        'Lower_Left_Z',
        'Box_Width',
        'Box_Length',
        'Box_Height',
        'Destination'
    ]
)

class Vehicle: # 다른 좌표계 사용
    X = 16
    Y = 28
    Z = 18
    total_volume = X*Y*Z*1000
    car_cost = 150000

    def __init__(self, datas:list[Data], name_to_index, OD_matrix):
        self.datas = datas
        self.name_to_index = name_to_index
        self.OD_matrix = OD_matrix

        self.used = [[[None]*self.Z for _ in range(self.Y)] for _ in range(self.X)]
        self.loaded_box_position_size = []
        self.shuffling_cost = 0
        self.travel_cost = 0
        self.route = []
        self.box_ids = []
        self.box_info:dict[str, Box] = dict()
        for data in self.datas:
            self.route.append(data.Destination)
            box_id = data.Box_ID
            self.box_ids.append(box_id)
            if box_id != None:
                position = (data.Lower_Left_X//10, data.Lower_Left_Y//10, data.Lower_Left_Z//10)
                size = (data.Box_Width//10, data.Box_Length//10, data.Box_Height//10)
                box = Box(*position, *size, data.Destination)
                self.box_info[box_id] = box
                self.load_box(box_id)
                self.loaded_box_position_size.append((position, size))
        self.loaded_box_position_size.reverse()

        # print(self.used)
        self.deliver()
        self.total_cost = self.car_cost + self.travel_cost + self.shuffling_cost

    def load_box(self, box_id):
        box = self.box_info[box_id]
        x, y, z = box.Lower_Left_X, box.Lower_Left_Y, box.Lower_Left_Z
        size_x, size_y, size_z = box.Box_Width, box.Box_Length, box.Box_Height
        for dx in range(size_x):
            for dy in range(size_y):
                for dz in range(size_z):
                    assert self.used[x+dx][y+dy][z+dz] == None
                    self.used[x+dx][y+dy][z+dz] = box_id

    def unload_box(self, box_id):
        moved = {box_id}

        def move_box(box_id, exclude):
            nonlocal moved
            box = self.box_info[box_id]
            x, y, z = box.Lower_Left_X, box.Lower_Left_Y, box.Lower_Left_Z
            size_x, size_y, size_z = box.Box_Width, box.Box_Length, box.Box_Height
            for target_z in range(z+size_z, self.Z):
                for dx in range(size_x):
                    for dy in range(size_y):
                        other_box_id = self.used[x+dx][y+dy][target_z]
                        if other_box_id == None: continue

                        other_box = self.box_info[other_box_id]
                        if other_box_id not in moved:
                            if other_box.Destination == exclude: continue
                            self.shuffling_cost += 500
                            moved.add(other_box_id)
                            move_box(other_box_id, exclude)
            for target_y in range(y-1, -1, -1):
                for dx in range(size_x):
                    for dz in range(size_z):
                        other_box_id = self.used[x+dx][target_y][z+dz]
                        if other_box_id == None: continue

                        other_box = self.box_info[other_box_id]
                        if other_box_id not in moved:
                            if other_box.Destination == exclude: continue
                            self.shuffling_cost += 500
                            moved.add(other_box_id)
                            move_box(other_box_id, exclude)

        # self.shuffling_cost += 500
        box = self.box_info[box_id]
        x, y, z = box.Lower_Left_X, box.Lower_Left_Y, box.Lower_Left_Z
        size_x, size_y, size_z = box.Box_Width, box.Box_Length, box.Box_Height
        move_box(box_id, box.Destination)
        for dx in range(size_x):
            for dy in range(size_y):
                for dz in range(size_z):
                    self.used[x+dx][y+dy][z+dz] = None

    def deliver(self):
        length = len(self.route)
        self.dist = 0
        for i in range(length-1):
            start = self.name_to_index[self.route[i]]
            end = self.name_to_index[self.route[i+1]]
            self.dist += self.OD_matrix[start][end]
            box_id = self.box_ids[i]
            if box_id != None: self.unload_box(box_id)
        self.travel_cost = self.dist * 0.5


def judge(data_file_name, distance_file_name):
    destinations, name_to_index, index_to_name = main.read_map(data_file_name)
    n = len(destinations)
    OD_matrix = main.read_OD_matrix(n, name_to_index, distance_file_name)
    orders = main.read_orders(n, name_to_index, data_file_name)

    def check_excel_error(datas: list[list[Data]]):
        # 행 검사
        order_num_data: dict[int, main.Order] = dict()
        for i in range(n):
            for order in orders[i]:
                order_num_data[order.order_num] = order

        for index, data in enumerate(datas):
            assert data[0].Destination == 'Depot', f'Vehicle {index}, route order 1 must be depot'
            assert data[-1].Destination == 'Depot', f'Vehicle {index}, last route order must be depot'
            for route_order, d in enumerate(data[1:-1], 2):
                assert d.Route_Order == route_order, f'Vehicle {index}, route order {route_order} : route order error'
                assert d.Stacking_Order == len(data) - route_order, f'Vehicle {index}, route order {route_order} : stacking order error'
                order_data = order_num_data[d.Order_Number]
                assert all([
                    order_data.destination == d.Destination,
                    order_data.box_id == d.Box_ID,
                    order_data.info == [100, 120, 160].index(d.Box_Width + d.Box_Length + d.Box_Height)
                ]), f'Vehicle {index}, route order {route_order} : data error'

        # 빠진 데이터 검사
        visited = [len(orders[i]) for i in range(n)]
        for index, data in enumerate(datas):
            for d in data[1:-1]:
                visited[name_to_index[d.Destination]] -= 1

        assert all(destination == 0 for destination in visited), 'There are destinations that have not been visited'

    # result_file_name = 'assignment1/sample Result.xlsx'
    result_file_name = 'Result.xlsx'
    wb = openpyxl.load_workbook(result_file_name)
    wb.sheetnames
    ws = wb[wb.sheetnames[0]]

    datas = []
    for index, row in enumerate(ws.rows):
        if index == 0: continue
        data = Data(*map(lambda x: x.value, row))
        if len(datas) == data.Vehicle_ID:
            datas.append([])
        datas[data.Vehicle_ID].append(data)

    check_excel_error(datas)

    vehicles: list[Vehicle] = []
    for data in datas:
        vehicles.append(Vehicle(data, name_to_index, OD_matrix))

    total_cost = 0
    car_cost = 0
    travel_cost = 0
    shuffling_cost = 0
    for index, vehicle in enumerate(vehicles):
        total_cost += vehicle.total_cost
        car_cost += vehicle.car_cost
        travel_cost += vehicle.travel_cost
        shuffling_cost += vehicle.shuffling_cost
        if DEBUG:
            print(f'Vehicle {index}')
            print(vehicle.total_cost)
            print(vehicle.car_cost, vehicle.travel_cost, vehicle.shuffling_cost)
            print()

        loaded_box_position_size = []
        for (x, y, z), (size_x, size_y, size_z) in vehicle.loaded_box_position_size:
            position = (x, 28-(y+size_y), z)
            size = (size_x, size_y, size_z)
            loaded_box_position_size.append((position, size))
        # viwer = visualize.box_viewer_3d(loaded_box_position_size)
        # viwer.show()

    if DEBUG:
        print(f'total cost : {int(total_cost):,}')
        print(f'car cost : {int(car_cost):,}')
        print(f'travel cost : {int(travel_cost):,}')
        print(f'shuffling cost : {int(shuffling_cost):,}')

    return int(total_cost)

def run_test(command):
    start_t = time.time()
    ret = os.system(command)
    running_time = time.time() - start_t

    msg = ''
    if ret == 0:
        try:
            total_cost = judge(data_file_name, distance_file_name)
        except Exception as reason:
            total_cost = 0
            msg = str(reason)
            print(reason)
    else:
        msg = 'error raised in main'
        total_cost = 0

    return ret, running_time, total_cost, msg

if __name__ == "__main__":
    assert basename(os.getcwd()) == 'routing'

    mode = int(input('mode : '))
    result = []
    if mode == 1: # CJ에서 준 테스트 데이터 파일
        data_file_name = 'Data_Set.json'
        distance_file_name = 'distance-data.txt'

        ret, running_time, total_cost, msg = run_test(f'python311 main.py {data_file_name} {distance_file_name}')

        if DEBUG:
            print(f'status : {ret}')
            print(f'running time : {running_time}')
            print(f'cost : {total_cost}')
            print()

        result.append(((ret, running_time, total_cost, msg),))
    elif mode == 2: # data 폴더 안에 들어있는 파일
        num = len(os.listdir('data')) // 2
        for i in range(1, num + 1):
            print(f'testing {i}')
            data_file_name = f'data/additional_data_{i:02d}.json'
            distance_file_name = f'data/additional_distance_data_{i:02d}.txt'
            assert os.path.exists(data_file_name)
            assert os.path.exists(distance_file_name)

            destinations, name_to_index, index_to_name = main.read_map(data_file_name)
            ret, running_time, total_cost, msg = run_test(f'python311 main.py {data_file_name} {distance_file_name}')

            if DEBUG:
                print(f'status : {ret}')
                print(f'running time : {running_time}')
                print(f'cost : {total_cost}')
                print()

            result.append(((ret, running_time, total_cost, msg),))
            if VISUALIZE and ret == 0: visualize.plot_vrp(data_file_name, distance_file_name)
    elif mode == 3: # 0.9로 싣는 알고리즘, 0.8로 싣는 알고리즘 비교
        num = len(os.listdir('data')) // 2
        for i in range(1, num + 1):
            print(f'testing {i}')
            data_file_name = f'data/additional_data_{i:02d}.json'
            distance_file_name = f'data/additional_distance_data_{i:02d}.txt'
            assert os.path.exists(data_file_name)
            assert os.path.exists(distance_file_name)

            ret1, running_time1, total_cost1, msg1 = run_test(f'python311 main.py {data_file_name} {distance_file_name}')
            ret2, running_time2, total_cost2, msg2 = run_test(f'python311 prev_main.py {data_file_name} {distance_file_name}')

            result.append(((ret1, running_time1, total_cost1, msg1), (ret2, running_time2, total_cost2, msg2)))

    n = len(result[0])
    data = []
    for i in range(n):
        for j, d in enumerate(result):
            status = d[i][0]
            if status != 0:
                print(f'Failed in test case {j+1} : {d[i][3]}')
    visualize.benchmark(result)